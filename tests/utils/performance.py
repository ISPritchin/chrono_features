# ruff: noqa: C901, PLR0912, PLR0915, ISC003, T201
import time
import platform
import psutil
from pathlib import Path

import numpy as np
import pandas as pd
import polars as pl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from rich.progress import (
    Progress,
    TextColumn,
    BarColumn,
    TaskProgressColumn,
    TimeElapsedColumn,
    TimeRemainingColumn,
    SpinnerColumn,
)
from rich.console import Console
from rich.panel import Panel
from rich.table import Table

from chrono_features import TSDataset, WindowType
from chrono_features.features._base import FeatureGenerator


def create_dataset(n_ids: int, n_timestamps: int) -> TSDataset:
    """Create a dataset with specified number of IDs and timestamps per ID."""
    ids = np.repeat(range(n_ids), n_timestamps)
    timestamps = np.tile(np.arange(1, n_timestamps + 1), n_ids)
    values = np.random.rand(n_ids * n_timestamps)
    data = pl.DataFrame(
        {
            "id": ids,
            "timestamp": timestamps,
            "value": values,
        },
    )
    return TSDataset(data, id_column_name="id", ts_column_name="timestamp")


def create_dataset_with_dynamic_windows(n_ids, n_timestamps, max_window_size=1000):
    """Create a dataset with an additional column for dynamic window lengths."""
    dataset = create_dataset(n_ids=n_ids, n_timestamps=n_timestamps)

    # Add a single dynamic window length column
    total_rows = len(dataset.data)

    # Values from 0 to max_window_size
    dynamic_len = np.random.randint(0, max_window_size + 1, size=total_rows)
    dataset.add_feature("dynamic_len", dynamic_len)

    return dataset


# Add to imports


def compare_performance(
    datasets: list[tuple[TSDataset, str]],
    implementations: list[tuple[type[FeatureGenerator], str]],
    window_types: list[WindowType],
    column_name: str = "value",
    output_file: str | None = None,
    time_threshold_seconds=5.0,
):
    """Compare performance of multiple implementations across different window types and datasets."""
    all_results = []
    console = Console()

    # Get system info
    system_info = {
        "Test date": time.strftime("%Y-%m-%d %H:%M:%S"),
        "OS": platform.system() + " " + platform.release(),
        "Python Version": platform.python_version(),
        "CPU": platform.processor(),
        "CPU Cores": psutil.cpu_count(logical=False),
        "CPU Threads": psutil.cpu_count(logical=True),
        "Total RAM (GB)": round(psutil.virtual_memory().total / (1024**3), 2),
    }

    # Display system info in a nice panel
    system_table = Table(title="System Information")
    system_table.add_column("Property", style="cyan")
    system_table.add_column("Value", style="green")

    for key, value in system_info.items():
        system_table.add_row(key, str(value))

    console.print("")
    console.print(Panel(system_table, title="Performance Test Environment", border_style="blue"))

    # Sort datasets by number of rows (from smallest to largest)
    datasets = sorted(datasets, key=lambda x: len(x[0].data))

    # Create a table to display dataset characteristics
    dataset_table = Table(title="Dataset Characteristics")
    dataset_table.add_column("Dataset Name", style="cyan")
    dataset_table.add_column("Unique IDs", style="magenta")
    dataset_table.add_column("Timestamps per ID", style="yellow")
    dataset_table.add_column("Total Rows", style="green")
    dataset_table.add_column("Max Dynamic Length", style="red")

    # Process each dataset and collect additional info
    dataset_additional_info = {}
    for dataset, dataset_name in datasets:
        # Get basic dataset info
        unique_ids = dataset.data[dataset.id_column_name].n_unique()
        timestamps_per_id = len(dataset.data) // unique_ids
        total_rows = len(dataset.data)

        # Check if dynamic_len column exists and get max value
        max_dynamic_len = "N/A"
        if "dynamic_len" in dataset.data.columns:
            max_dynamic_len = str(dataset.data["dynamic_len"].max())
            if dataset_name not in dataset_additional_info:
                dataset_additional_info[dataset_name] = {}
            dataset_additional_info[dataset_name]["Max dynamic_len"] = max_dynamic_len

        # Add row to the table
        dataset_table.add_row(dataset_name, str(unique_ids), str(timestamps_per_id), str(total_rows), max_dynamic_len)

    # Display the dataset characteristics table
    console.print(Panel(dataset_table, title="Datasets to be Tested", border_style="green"))

    # Display window types
    window_table = Table(title="Window Types")
    window_table.add_column("Window Type", style="blue")
    window_table.add_column("Description", style="yellow")

    for window_type in window_types:
        window_table.add_row(str(window_type), window_type._description)

    console.print(Panel(window_table, title="Window Types to be Tested", border_style="magenta"))

    # Dictionary to track implementations that exceeded the time threshold
    # Key: (implementation_name, window_type_str), Value: Boolean (exceeded threshold)
    exceeded_threshold = {}

    # Calculate total number of tests to run
    total_tests = len(datasets) * len(window_types) * len(implementations)

    # Create rich progress bar
    with Progress(
        SpinnerColumn(),
        TextColumn("[bold blue]{task.description}"),
        BarColumn(bar_width=40),
        TaskProgressColumn(),
        TextColumn("•"),
        TimeElapsedColumn(),
        TextColumn("•"),
        TimeRemainingColumn(),
    ) as progress:
        # Add overall task
        overall_task = progress.add_task("[yellow]Overall Progress", total=total_tests)

        # Process each window type first
        for window_type in window_types:
            window_info = str(window_type)

            # Add window type task
            window_task = progress.add_task(f"[green]Window: {window_info}", total=len(implementations) * len(datasets))

            # Process each implementation
            for impl_class, name in implementations:
                # Add implementation task
                impl_task = progress.add_task(f"[magenta]Implementation: {name}", total=len(datasets))

                # Process each dataset (from smallest to largest)
                for dataset_idx, (dataset, dataset_name) in enumerate(datasets):
                    # Update task description
                    current_test = f"[cyan]Testing: {name} on {window_info} with {dataset_name}"
                    progress.update(overall_task, description=current_test)

                    # Check if this implementation already exceeded threshold for this window type
                    if exceeded_threshold.get((name, window_info), False):
                        console.print(
                            f"[yellow]Skipping {name} with {window_info} on {dataset_name} "
                            f"(exceeded time threshold on smaller dataset)",
                        )

                        # Create result dictionary with skipped status
                        all_results.append(
                            {
                                "dataset_name": dataset_name,
                                "window_type": window_info,
                                "implementation": f"{name} ({impl_class.__name__})",
                                "execution_time": "SKIPPED (time threshold exceeded)",
                                "dataset_info": {
                                    "Number of unique IDs": dataset.data[dataset.id_column_name].n_unique(),
                                    "Timestamps per ID": len(dataset.data)
                                    // dataset.data[dataset.id_column_name].n_unique(),
                                    "Total rows": len(dataset.data),
                                },
                            },
                        )
                        # Update progress bars
                        progress.update(overall_task, advance=1)
                        progress.update(window_task, advance=1)
                        progress.update(impl_task, advance=1)
                        continue

                    # Create transformer
                    transformer = impl_class(
                        columns=column_name,
                        window_types=window_type,
                        out_column_names=f"{name}_result",
                    )

                    # Measure execution time
                    start_time = time.time()
                    try:
                        _ = transformer.transform(dataset.clone())
                        execution_time = time.time() - start_time

                        # Check if execution time exceeded threshold and we're not on the last dataset
                        if execution_time > time_threshold_seconds and dataset_idx < len(datasets) - 1:
                            console.print(
                                f"[red]Warning: {name} with {window_info} on {dataset_name} "
                                f"exceeded time threshold ({execution_time:.2f}s > {time_threshold_seconds:.2f}s)",
                            )
                            # Mark this implementation as exceeding threshold for this window type
                            exceeded_threshold[(name, window_info)] = True

                        # Create result dictionary
                        result = {
                            "dataset_name": dataset_name,
                            "window_type": window_info,
                            "implementation": f"{name} ({impl_class.__name__})",
                            "execution_time": execution_time,
                            "dataset_info": {
                                "Number of unique IDs": dataset.data[dataset.id_column_name].n_unique(),
                                "Timestamps per ID": len(dataset.data)
                                // dataset.data[dataset.id_column_name].n_unique(),
                                "Total rows": len(dataset.data),
                            },
                        }
                        all_results.append(result)

                    except Exception as e:  # noqa: BLE001
                        console.print(f"[bold red]Error with {name} on {window_info} for {dataset_name}: {e!s}")
                        # Add result with error
                        all_results.append(
                            {
                                "dataset_name": dataset_name,
                                "window_type": window_info,
                                "implementation": f"{name} ({impl_class.__name__})",
                                "execution_time": f"error: {e!s}",
                                "dataset_info": {
                                    "Number of unique IDs": dataset.data[dataset.id_column_name].n_unique(),
                                    "Timestamps per ID": len(dataset.data)
                                    // dataset.data[dataset.id_column_name].n_unique(),
                                    "Total rows": len(dataset.data),
                                },
                            },
                        )

                    # Update progress bars
                    progress.update(overall_task, advance=1)
                    progress.update(window_task, advance=1)
                    progress.update(impl_task, advance=1)

                # Complete implementation task
                progress.update(impl_task, completed=len(datasets))

            # Complete window task
            progress.update(window_task, completed=len(implementations) * len(datasets))

    # Save to Excel file if requested
    if output_file:
        # Create a new workbook or load existing
        if Path(output_file).exists():
            wb = load_workbook(output_file)
        else:
            wb = Workbook()
            # Remove default sheet if it exists
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        # Determine sheet name based on implementations
        # Extract base class name without "WithOptimization" or "WithoutOptimization"
        base_class_names = set()
        for impl_class, _ in implementations:
            class_name = impl_class.__name__
            if "WithOptimization" in class_name:
                base_name = class_name.replace("WithOptimization", "")
                base_class_names.add(base_name)
            elif "WithoutOptimization" in class_name:
                base_name = class_name.replace("WithoutOptimization", "")
                base_class_names.add(base_name)
            else:
                base_class_names.add(class_name)

        sheet_name = ", ".join(sorted(base_class_names))

        # Create or get sheet
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]  # Replace existing sheet
        sheet = wb.create_sheet(title=sheet_name)

        # Add title for system info
        sheet["B1"] = "Stend description"
        sheet["B1"].font = Font(bold=True)

        last_col = get_column_letter(len(datasets) + 2)
        sheet.merge_cells(f"B1:{last_col}1")
        sheet["B1"].alignment = Alignment(horizontal="center")

        # Add system info
        row = 3
        for key, value in system_info.items():
            sheet[f"B{row}"] = key
            sheet[f"C{row}"] = str(value)
            sheet[f"B{row}"].font = Font(italic=True)
            row += 1

        # Add empty row
        row += 1

        # Add dataset info header
        sheet[f"B{row}"] = "Dataset Information"
        sheet[f"B{row}"].font = Font(bold=True)
        sheet.merge_cells(f"B{row}:{last_col}{row}")
        sheet[f"B{row}"].alignment = Alignment(horizontal="center")
        row += 1

        # Get dataset names in the original order
        dataset_names = [name for _, name in datasets]

        # Add dataset info columns
        for i, dataset_name in enumerate(dataset_names):
            col = get_column_letter(i + 3)
            sheet[f"{col}{row}"] = dataset_name
            sheet[f"{col}{row}"].font = Font(bold=True)
            sheet[f"{col}{row}"].alignment = Alignment(horizontal="center")
        row += 1

        # Add dataset details
        dataset_metrics = ["Number of unique IDs", "Timestamps per ID", "Total rows"]

        # Add max dynamic_len metric if it exists for any dataset
        if any("Max dynamic_len" in dataset_additional_info.get(name, {}) for _, name in datasets):
            dataset_metrics.append("Max dynamic_len")

        for metric in dataset_metrics:
            sheet[f"B{row}"] = metric

            for i, dataset_name in enumerate(dataset_names):
                col = get_column_letter(i + 3)

                if metric in ["Number of unique IDs", "Timestamps per ID", "Total rows"]:
                    # Find first result for this dataset
                    dataset_result = next((r for r in all_results if r["dataset_name"] == dataset_name), None)
                    if dataset_result and metric in dataset_result["dataset_info"]:
                        sheet[f"{col}{row}"] = dataset_result["dataset_info"][metric]
                        # Center align the values
                        sheet[f"{col}{row}"].alignment = Alignment(horizontal="center")
                elif (
                    metric == "Max dynamic_len"
                    and dataset_name in dataset_additional_info
                    and "Max dynamic_len" in dataset_additional_info[dataset_name]
                ):
                    sheet[f"{col}{row}"] = dataset_additional_info[dataset_name]["Max dynamic_len"]
                    sheet[f"{col}{row}"].alignment = Alignment(horizontal="center")
            row += 1

        # Add empty row
        row += 1

        # Add performance results header
        sheet[f"A{row}"] = "Performance Results"
        sheet[f"A{row}"].font = Font(bold=True)

        last_perf_col = get_column_letter(len(dataset_names) + 2)
        sheet.merge_cells(f"A{row}:{last_perf_col}{row}")
        sheet[f"A{row}"].alignment = Alignment(horizontal="center")
        row += 1

        # Add column headers for performance results
        sheet[f"A{row}"] = "Window"
        sheet[f"B{row}"] = "Implementation"
        sheet[f"A{row}"].font = Font(bold=True)
        sheet[f"B{row}"].font = Font(bold=True)

        for i in range(len(dataset_names)):
            col = get_column_letter(i + 3)
            sheet[f"{col}{row}"] = "Time (s)"
            sheet[f"{col}{row}"].font = Font(bold=True)
            sheet[f"{col}{row}"].alignment = Alignment(horizontal="center")
        row += 1

        # Add empty row
        row += 1

        # Group results by window type
        window_types_in_results = sorted({r["window_type"] for r in all_results})

        for window_type in window_types_in_results:
            # Add window type as header
            sheet[f"A{row}"] = window_type
            sheet[f"A{row}"].font = Font(bold=True)

            # Get all implementations for this window type
            implementations_for_window = sorted(
                {r["implementation"] for r in all_results if r["window_type"] == window_type},
            )

            # Add results for each implementation
            for i, implementation in enumerate(implementations_for_window):
                if i > 0:  # Only set implementation name for first row
                    row += 1
                    sheet[f"A{row}"] = ""  # Empty cell for window type

                sheet[f"B{row}"] = implementation

                # Add execution times for each dataset
                for j, dataset_name in enumerate(dataset_names):
                    # Find result for this implementation, window type, and dataset
                    result = next(
                        (
                            r
                            for r in all_results
                            if r["window_type"] == window_type
                            and r["implementation"] == implementation
                            and r["dataset_name"] == dataset_name
                        ),
                        None,
                    )

                    if result:
                        col = get_column_letter(j + 3)
                        time_value = result["execution_time"]
                        if isinstance(time_value, (int, float)):
                            sheet[f"{col}{row}"] = f"{time_value:.6f}"
                        else:
                            sheet[f"{col}{row}"] = str(time_value)
                        # Center align the values
                        sheet[f"{col}{row}"].alignment = Alignment(horizontal="center")

            # Add two empty rows between window types
            row += 3

        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)

            # Set minimum and maximum widths
            adjusted_width = min(max(max_length + 2, 15), 60)  # Min 15, Max 60
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(output_file)
        print(f"Results for {sheet_name} saved to Excel file: {output_file}")

    return pd.DataFrame(all_results)
